# excel_salida.py

from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import (
    BarChart,
    LineChart,
    PieChart,
    Reference,
)
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import (
    Table,
    TableStyleInfo,
)


COLOR_AZUL = "4472C4"
COLOR_AZUL_OSCURO = "203864"
COLOR_AMARILLO = "FFF2CC"
COLOR_VERDE = "E2F0D9"


# ============================================================
# FORMATO GENERAL
# ============================================================

def limpiar_nombre_tabla(
    nombre: str,
) -> str:
    """
    Devuelve un nombre válido y único para una tabla de Excel.
    """
    nombre_limpio = re.sub(
        r"[^A-Za-z0-9_]",
        "",
        nombre,
    )

    if not nombre_limpio:
        nombre_limpio = "TablaDatos"

    if nombre_limpio[0].isdigit():
        nombre_limpio = (
            "Tabla_" + nombre_limpio
        )

    return nombre_limpio[:200]


def ajustar_columnas(
    hoja,
) -> None:
    """
    Ajusta el ancho de las columnas.
    """
    for columna in hoja.columns:
        letra = get_column_letter(
            columna[0].column
        )

        ancho_maximo = 0

        for celda in columna:
            valor = (
                ""
                if celda.value is None
                else str(celda.value)
            )

            ancho_maximo = max(
                ancho_maximo,
                len(valor),
            )

        hoja.column_dimensions[
            letra
        ].width = min(
            ancho_maximo + 3,
            55,
        )


def aplicar_estilo_encabezado(
    hoja,
) -> None:
    """
    Da formato al encabezado.
    """
    relleno = PatternFill(
        fill_type="solid",
        fgColor=COLOR_AZUL,
    )

    for celda in hoja[1]:
        celda.fill = relleno
        celda.font = Font(
            bold=True,
            color="FFFFFF",
        )
        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    hoja.row_dimensions[1].height = 32


def convertir_en_tabla(
    hoja,
    nombre_tabla: str,
) -> None:
    """
    Convierte el rango en tabla.

    Importante:
    no agrega AutoFilter adicional, porque la tabla
    ya incluye el filtro y evita reparaciones de Excel.
    """
    if hoja.max_row < 2:
        return

    ultima_columna = get_column_letter(
        hoja.max_column
    )

    referencia = (
        f"A1:{ultima_columna}"
        f"{hoja.max_row}"
    )

    tabla = Table(
        displayName=limpiar_nombre_tabla(
            nombre_tabla
        ),
        ref=referencia,
    )

    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    hoja.add_table(tabla)


def aplicar_formato_fechas(
    hoja,
) -> None:
    """
    Aplica formato según el encabezado.
    """
    encabezados = {
        celda.value: celda.column
        for celda in hoja[1]
    }

    for nombre in [
        "Inicio original",
        "Inicio",
        "Fin",
    ]:
        if nombre in encabezados:
            columna = encabezados[nombre]

            for fila in range(
                2,
                hoja.max_row + 1,
            ):
                hoja.cell(
                    row=fila,
                    column=columna,
                ).number_format = (
                    "dd/mm/yyyy hh:mm:ss"
                )

    if "Fecha capacitación" in encabezados:
        columna = encabezados[
            "Fecha capacitación"
        ]

        for fila in range(
            2,
            hoja.max_row + 1,
        ):
            hoja.cell(
                row=fila,
                column=columna,
            ).number_format = "dd/mm/yyyy"


def destacar_estados(
    hoja,
) -> None:
    """
    Marca REVISAR en amarillo y OK en verde.
    """
    encabezados = {
        celda.value: celda.column
        for celda in hoja[1]
    }

    if "Estado" not in encabezados:
        return

    columna = encabezados["Estado"]

    relleno_revisar = PatternFill(
        fill_type="solid",
        fgColor=COLOR_AMARILLO,
    )

    relleno_ok = PatternFill(
        fill_type="solid",
        fgColor=COLOR_VERDE,
    )

    for fila in range(
        2,
        hoja.max_row + 1,
    ):
        celda = hoja.cell(
            row=fila,
            column=columna,
        )

        if celda.value == "REVISAR":
            celda.fill = relleno_revisar
            celda.font = Font(
                bold=True,
                color="9C6500",
            )

        elif celda.value == "OK":
            celda.fill = relleno_ok
            celda.font = Font(
                bold=True,
                color="375623",
            )


# ============================================================
# PREPARACIÓN DE DATOS
# ============================================================

def preparar_base_corregida(
    resultado: pd.DataFrame,
) -> pd.DataFrame:
    """
    Ordena las columnas de la salida principal.
    """
    columnas = [
        "ID",
        "Inicio original",
        "Inicio",
        "Fin",
        "Apellido",
        "Nombre",
        "DNI",
        "Fecha capacitación",
        "Complejo",
        "Cargo",
        "Tipo personal",
        "Tipo personal corregido",
        "Rol asignado",
        "Módulos originales",
        "Módulos esperados",
        "Módulos corregidos",
        "Cantidad de registros",
        "Fecha corregida",
        "Regla aplicada",
        "Corrección realizada",
        "Estado",
    ]

    columnas_existentes = [
        columna
        for columna in columnas
        if columna in resultado.columns
    ]

    return resultado[
        columnas_existentes
    ].copy()


def crear_resumen(
    resultado: pd.DataFrame,
) -> pd.DataFrame:
    """
    Crea indicadores principales.
    """
    return pd.DataFrame(
        {
            "Indicador": [
                "Usuarios únicos",
                "Duplicados consolidados",
                "Fechas corregidas",
                "Pendientes de revisión",
                "Emisor",
                "Solicitante",
                "Verificador",
                "Ejecutante",
                "Formación de formadores",
            ],
            "Cantidad": [
                len(resultado),
                int(
                    (
                        resultado[
                            "Cantidad de registros"
                        ] > 1
                    ).sum()
                ),
                int(
                    (
                        resultado[
                            "Fecha corregida"
                        ] == "Sí"
                    ).sum()
                ),
                int(
                    (
                        resultado["Estado"]
                        == "REVISAR"
                    ).sum()
                ),
                int(
                    (
                        resultado["Rol asignado"]
                        == "Emisor"
                    ).sum()
                ),
                int(
                    (
                        resultado["Rol asignado"]
                        == "Solicitante"
                    ).sum()
                ),
                int(
                    (
                        resultado["Rol asignado"]
                        == "Verificador"
                    ).sum()
                ),
                int(
                    (
                        resultado["Rol asignado"]
                        == "Ejecutante"
                    ).sum()
                ),
                int(
                    (
                        resultado["Rol asignado"]
                        == "Formación de formadores"
                    ).sum()
                ),
            ],
        }
    )


# ============================================================
# GENERACIÓN DEL EXCEL
# ============================================================

def generar_excel(
    resultado: pd.DataFrame,
    ruta_salida: str | Path,
) -> Path:
    """
    Genera el Excel final:

    - Base corregida
    - Correcciones realizadas
    - Pendientes
    - Resumen
    - Usuarios por rol
    - Usuarios por cargo
    - YPF y contratistas
    - Por fecha
    """
    ruta_salida = Path(ruta_salida)

    ruta_salida.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    base_corregida = preparar_base_corregida(
        resultado
    )

    correcciones = resultado[
        resultado["Corrección realizada"]
        != "Sin cambios"
    ].copy()

    pendientes = resultado[
        resultado["Estado"] == "REVISAR"
    ].copy()

    resumen = crear_resumen(
        resultado
    )

    usuarios_por_rol = (
        resultado.groupby(
            "Rol asignado",
            dropna=False,
        )
        .size()
        .reset_index(name="Cantidad")
        .sort_values(
            "Cantidad",
            ascending=False,
        )
    )

    usuarios_por_cargo = (
        resultado.groupby(
            "Cargo",
            dropna=False,
        )
        .size()
        .reset_index(name="Cantidad")
        .sort_values(
            "Cantidad",
            ascending=False,
        )
    )

    usuarios_por_tipo = (
    resultado.groupby(
        "Tipo personal corregido",
        dropna=False,
    )
    .size()
    .reset_index(name="Cantidad")
    .rename(
        columns={
            "Tipo personal corregido": "Tipo personal"
        }
    )
    .sort_values(
        "Cantidad",
        ascending=False,
    )
    )

    datos_fecha = resultado.dropna(
        subset=["Fecha capacitación"]
    ).copy()

    datos_fecha["Fecha"] = (
        datos_fecha[
            "Fecha capacitación"
        ].dt.date
    )

    capacitaciones_por_fecha = (
        datos_fecha.groupby("Fecha")
        .size()
        .reset_index(name="Cantidad")
        .sort_values("Fecha")
    )

    with pd.ExcelWriter(
        ruta_salida,
        engine="openpyxl",
    ) as writer:

        base_corregida.to_excel(
            writer,
            sheet_name="Base corregida",
            index=False,
        )

        correcciones.to_excel(
            writer,
            sheet_name="Correcciones realizadas",
            index=False,
        )

        pendientes.to_excel(
            writer,
            sheet_name="Pendientes",
            index=False,
        )

        resumen.to_excel(
            writer,
            sheet_name="Resumen",
            index=False,
        )

        usuarios_por_rol.to_excel(
            writer,
            sheet_name="Usuarios por rol",
            index=False,
        )

        usuarios_por_cargo.to_excel(
            writer,
            sheet_name="Usuarios por cargo",
            index=False,
        )

        usuarios_por_tipo.to_excel(
            writer,
            sheet_name="YPF y contratistas",
            index=False,
        )

        capacitaciones_por_fecha.to_excel(
            writer,
            sheet_name="Por fecha",
            index=False,
        )

    libro = load_workbook(
        ruta_salida
    )

    nombres_tablas = {
        "Base corregida": (
            "TablaBaseCorregida"
        ),
        "Correcciones realizadas": (
            "TablaCorrecciones"
        ),
        "Pendientes": (
            "TablaPendientes"
        ),
        "Resumen": (
            "TablaResumen"
        ),
        "Usuarios por rol": (
            "TablaUsuariosRol"
        ),
        "Usuarios por cargo": (
            "TablaUsuariosCargo"
        ),
        "YPF y contratistas": (
            "TablaTipoPersonal"
        ),
        "Por fecha": (
            "TablaPorFecha"
        ),
    }

    for hoja in libro.worksheets:
        aplicar_estilo_encabezado(
            hoja
        )

        ajustar_columnas(
            hoja
        )

        hoja.freeze_panes = "A2"

        # No usar hoja.auto_filter.ref aquí.
        convertir_en_tabla(
            hoja,
            nombres_tablas[hoja.title],
        )

        aplicar_formato_fechas(
            hoja
        )

        destacar_estados(
            hoja
        )

        for fila in hoja.iter_rows(
            min_row=2,
        ):
            for celda in fila:
                celda.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

    # --------------------------------------------------------
    # Gráfico de usuarios por rol
    # --------------------------------------------------------
    hoja_roles = libro[
        "Usuarios por rol"
    ]

    if hoja_roles.max_row > 1:
        grafico_roles = BarChart()

        grafico_roles.title = (
            "Usuarios por rol"
        )
        grafico_roles.y_axis.title = (
            "Cantidad"
        )
        grafico_roles.x_axis.title = (
            "Rol"
        )
        grafico_roles.height = 8
        grafico_roles.width = 15

        datos = Reference(
            hoja_roles,
            min_col=2,
            min_row=1,
            max_row=hoja_roles.max_row,
        )

        categorias = Reference(
            hoja_roles,
            min_col=1,
            min_row=2,
            max_row=hoja_roles.max_row,
        )

        grafico_roles.add_data(
            datos,
            titles_from_data=True,
        )

        grafico_roles.set_categories(
            categorias
        )

        hoja_roles.add_chart(
            grafico_roles,
            "D2",
        )

    # --------------------------------------------------------
    # Gráfico por fecha
    # --------------------------------------------------------
    hoja_fecha = libro[
        "Por fecha"
    ]

    if hoja_fecha.max_row > 1:
        grafico_fecha = LineChart()

        grafico_fecha.title = (
            "Capacitaciones por fecha"
        )
        grafico_fecha.y_axis.title = (
            "Usuarios"
        )
        grafico_fecha.x_axis.title = (
            "Fecha"
        )
        grafico_fecha.height = 8
        grafico_fecha.width = 16

        datos = Reference(
            hoja_fecha,
            min_col=2,
            min_row=1,
            max_row=hoja_fecha.max_row,
        )

        categorias = Reference(
            hoja_fecha,
            min_col=1,
            min_row=2,
            max_row=hoja_fecha.max_row,
        )

        grafico_fecha.add_data(
            datos,
            titles_from_data=True,
        )

        grafico_fecha.set_categories(
            categorias
        )

        hoja_fecha.add_chart(
            grafico_fecha,
            "D2",
        )

    # --------------------------------------------------------
    # Gráfico YPF vs. contratistas
    # --------------------------------------------------------
    hoja_tipo = libro[
        "YPF y contratistas"
    ]

    if hoja_tipo.max_row > 1:
        grafico_tipo = PieChart()

        grafico_tipo.title = (
            "Distribución por tipo de personal"
        )
        grafico_tipo.height = 8
        grafico_tipo.width = 12

        datos = Reference(
            hoja_tipo,
            min_col=2,
            min_row=1,
            max_row=hoja_tipo.max_row,
        )

        categorias = Reference(
            hoja_tipo,
            min_col=1,
            min_row=2,
            max_row=hoja_tipo.max_row,
        )

        grafico_tipo.add_data(
            datos,
            titles_from_data=True,
        )

        grafico_tipo.set_categories(
            categorias
        )

        hoja_tipo.add_chart(
            grafico_tipo,
            "D2",
        )

    libro.save(
        ruta_salida
    )

    return ruta_salida